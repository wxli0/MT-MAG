import sys
import os
import pandas as pd 
import openpyxl
import numpy as np
from math import exp
import openpyxl

""" 
Add 'max' column to a sheet in a file to represent the maximum posterior \
    likelihhods for all classes per classification

    Command line arguments:
    :param argv[1]: file_path, file path
    :type argv[1]: str
    :param argv[2]: sheet, sheet name
    :type argv[2]: str
    :Example: python3 add_max.py outputs-r202/g__Prevotella.xlsx g__Prevotella-t-p
"""

file_path = sys.argv[1]
sheet = sys.argv[2]

df = pd.read_excel(file_path, index_col=0, header=0, sheet_name=sheet)
pred = df['prediction']
if 'max' in df.columns:
    del df['max']
if 'prediction' in df.columns:
    del df['prediction']
if 'rejection' in df.columns:
    del df['rejection']
if 'actual' in df.columns:
    del df['actual']
if 'gtdb-Tk' in df.columns:
    del df['gtdb-Tk']
if 'rejection-f' in df.columns:
    del df['rejection-f']

df_softmax = pd.DataFrame(df.to_numpy(), columns=df.columns)
df_softmax['max'] = np.max(df_softmax, axis=1)
df_softmax.index = df.index
df_softmax['prediction'] = pred
print(df_softmax)

wb = openpyxl.load_workbook(file_path)
del wb[sheet]
if len(wb.sheetnames) != 0:
    wb.save(file_path)

mode='a'
if len(wb.sheetnames) == 0:
    mode='w'

print("mode is", mode)
with pd.ExcelWriter(file_path, engine="openpyxl", mode=mode) as writer:  
    df_softmax.to_excel(writer, sheet_name = sheet, index=True)
writer.save()
writer.close()



