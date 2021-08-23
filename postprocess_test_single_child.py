"""
Add 'rejection-f' column of the single child taxa to \
    the classification result file
:param sys.argv[1]: file. The file path of the classification result file.
:type sys.argv[1]: str
"""

import openpyxl
import os
import pandas as pd 
import sys


file = sys.argv[1]

# remove quadratic score sheet, fix prediction column
df = pd.read_excel(file, sheet_name = "quadratic-svm-score", index_col=0, header=0)

classes = df.columns.tolist()
for i in range(len(classes)-1):
    c = classes[i]
    index = classes.index(c)+1
    df = df.replace({'prediction': {index: "-".join(c.split('-')[1:])}})

with pd.ExcelWriter(file, engine="openpyxl", mode='a') as writer: 
    df.to_excel(writer, sheet_name = file.split('/')[-1][:-5]+"_pred-t-p", index=True)
    writer.save()
    writer.close()

wb = openpyxl.load_workbook(file)
del wb["quadratic-svm-score"]
wb.save(file)

# add rejection-f
rejection_f = []
sheet = file.split('/')[-1][:-5]+"_pred-t-p"
df_new = pd.read_excel(file, index_col=0, header=0, sheet_name=sheet) # read in the new sheet 
taxon = df_new.columns[0] # the only child taxon
for index, row in df.iterrows():
    if float(row[taxon]) < 0:
        rejection_f.append(row['prediction']+'(reject)')
    else:
        rejection_f.append(row['prediction'])
if 'rejection-f' in df.columns:
    del df['rejection-f']
df['rejection-f'] = rejection_f

wb = openpyxl.load_workbook(file)
del wb[sheet]

mode = 'w'
if len(wb.sheetnames) != 0:
    wb.save(file)
    mode = 'a'

with pd.ExcelWriter(file, engine="openpyxl", mode=mode) as writer:  
    df.to_excel(writer, sheet_name = sheet, index=True)
writer.save()
writer.close()

